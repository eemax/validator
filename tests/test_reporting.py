from pathlib import Path

from openpyxl import load_workbook

from centric_mdm_validation.centric.store import build_reconstruction_coverage_check
from centric_mdm_validation.io import read_json_records
from centric_mdm_validation.models import CentricProductPayload
from centric_mdm_validation.reporting import DppReadinessReporter, ReconstructionCheckReporter
from centric_mdm_validation.validation import DppReadinessValidator, DppRuleSet


def test_reporter_writes_expected_files(tmp_path: Path) -> None:
    rules = DppRuleSet.from_yaml(Path("tests/fixtures/dpp-readiness.yml"))
    records = read_json_records(Path("tests/fixtures/projected-products.json"))
    payloads = [CentricProductPayload.model_validate(record) for record in records]
    run = DppReadinessValidator(rules).validate_many(payloads)

    DppReadinessReporter().write_all(run, tmp_path)

    assert (tmp_path / "dpp-readiness-summary.md").exists()
    assert (tmp_path / "dpp-readiness-products.csv").exists()
    assert (tmp_path / "dpp-readiness-issues.csv").exists()
    assert (tmp_path / "dpp-readiness.xlsx").exists()

    summary = (tmp_path / "dpp-readiness-summary.md").read_text(encoding="utf-8")
    assert "66.67%" in summary
    assert "MATERIAL_COMPOSITION" in summary


def test_reconstruction_check_reporter_writes_single_summary_and_workbook(
    tmp_path: Path,
) -> None:
    run = build_reconstruction_coverage_check(
        {
            "styles": [
                {
                    "id": "S1",
                    "active_colorways": {"0": "C1"},
                    "product_sizes": {"0": "C2"},
                }
            ],
            "colorways": [],
            "sizes": [{"id": "C2"}],
        }
    )

    ReconstructionCheckReporter().write_all(run, tmp_path)

    assert (tmp_path / "reconstruction-check-summary.md").is_file()
    assert (tmp_path / "reconstruction-check.xlsx").is_file()
    assert not (tmp_path / "reconstruction-check-products.csv").exists()
    assert not (tmp_path / "reconstruction-check-issues.csv").exists()

    workbook = load_workbook(tmp_path / "reconstruction-check.xlsx", read_only=True)
    assert workbook.sheetnames == [
        "Summary",
        "Issues",
        "Unresolved Refs",
        "Endpoint Coverage",
        "Definitions",
    ]
    issue_rows = list(workbook["Issues"].iter_rows(values_only=True))
    assert issue_rows[0] == ("code", "severity", "bucket", "count")
    assert any(row[0] == "REFERENCED_RECORD_NOT_SEEN" for row in issue_rows[1:])
    unresolved_rows = list(workbook["Unresolved Refs"].iter_rows(values_only=True))
    assert unresolved_rows[0] == ("relationship", "target_endpoint", "status", "count")
    assert any(row[2] == "not_seen" for row in unresolved_rows[1:])
