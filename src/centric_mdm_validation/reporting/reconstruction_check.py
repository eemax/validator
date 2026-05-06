from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from centric_mdm_validation.models import ProductValidationResult, ValidationRunResult


class ReconstructionCheckReporter:
    def write_all(self, run: ValidationRunResult | dict[str, Any], output_dir: Path) -> None:
        output_dir.mkdir(parents=True, exist_ok=True)
        report = self._normalize_run(run)
        self.write_summary_markdown(report, output_dir / "reconstruction-check-summary.md")
        self.write_workbook(report, output_dir / "reconstruction-check.xlsx")

    def write_summary_markdown(self, report: dict[str, Any], path: Path) -> None:
        summary = report.get("summary", {})
        issue_rows = report.get("issue_counts", [])
        relationship_rows = report.get("relationship_coverage", [])
        endpoint_rows = report.get("endpoint_coverage", [])
        unresolved_rows = report.get("unresolved_refs", [])
        lines = [
            "# Reconstruction Check Summary",
            "",
            f"- Rule set: `{report.get('rule_set_version', '')}`",
            f"- Styles in store: `{summary.get('styles', 0)}`",
            f"- Endpoints in store: `{summary.get('endpoints', len(endpoint_rows))}`",
            (
                "- Relationships checked: "
                f"`{summary.get('relationships_checked', len(relationship_rows))}`"
            ),
            f"- Declared refs: `{summary.get('declared_refs', 0)}`",
            f"- Seen refs: `{summary.get('seen_refs', 0)}`",
            f"- Missing refs: `{summary.get('missing_refs', 0)}`",
            f"- Invalid refs: `{summary.get('invalid_refs', 0)}`",
            f"- Reference coverage: `{summary.get('coverage_percent', 0.0)}%`",
            "",
            "## Relationship Coverage",
            "",
            (
                "| Relationship | Source | Target | Declared | Seen | Missing | "
                "Invalid | Coverage | Source Records With Missing Refs |"
            ),
            "| --- | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |",
        ]
        lines.extend(
            (
                f"| {row.get('relationship', '')} | {row.get('source_endpoint', '')} | "
                f"{row.get('target_endpoint', '')} | {row.get('declared_refs', 0)} | "
                f"{row.get('seen_refs', 0)} | {row.get('missing_refs', 0)} | "
                f"{row.get('invalid_refs', 0)} | {row.get('coverage_percent', 0.0)}% | "
                f"{row.get('source_records_with_missing_refs', 0)} |"
            )
            for row in relationship_rows
        )
        lines.extend(
            [
                "",
                "## Issue Counts",
                "",
                "| Code | Severity | Bucket | Count |",
                "| --- | --- | --- | ---: |",
            ]
        )
        lines.extend(
            f"| {row.get('code', '')} | {row.get('severity', '')} | "
            f"{row.get('bucket', '')} | {row.get('count', 0)} |"
            for row in issue_rows
        )
        lines.extend(
            [
                "",
                "## Unresolved Ref Counts",
                "",
                "| Relationship | Target | Status | Count |",
                "| --- | --- | --- | ---: |",
            ]
        )
        lines.extend(
            f"| {row.get('relationship', '')} | {row.get('target_endpoint', '')} | "
            f"{row.get('status', '')} | {row.get('count', 0)} |"
            for row in unresolved_rows
        )
        lines.extend(
            [
                "",
                "## Endpoint Coverage",
                "",
                "| Endpoint | Records |",
                "| --- | ---: |",
            ]
        )
        lines.extend(
            f"| {row.get('endpoint', '')} | {row.get('records', 0)} |" for row in endpoint_rows
        )
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def write_workbook(self, report: dict[str, Any], path: Path) -> None:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        self._append_rows(summary, ["metric", "value"], self._summary_rows(report))

        issues = workbook.create_sheet("Issues")
        self._append_rows(
            issues,
            ["code", "severity", "bucket", "count"],
            report.get("issue_counts", []),
        )

        unresolved = workbook.create_sheet("Unresolved Refs")
        self._append_rows(
            unresolved,
            ["relationship", "target_endpoint", "status", "count"],
            report.get("unresolved_refs", []),
        )

        coverage = workbook.create_sheet("Endpoint Coverage")
        self._append_rows(
            coverage,
            [
                "relationship",
                "source_endpoint",
                "target_endpoint",
                "source_records",
                "source_records_with_refs",
                "source_records_with_missing_refs",
                "declared_refs",
                "seen_refs",
                "missing_refs",
                "invalid_refs",
                "coverage_percent",
            ],
            report.get("relationship_coverage", []),
        )
        if report.get("endpoint_coverage"):
            coverage.append([])
            coverage.append(["endpoint", "records"])
            for row in report.get("endpoint_coverage", []):
                coverage.append([row.get("endpoint"), row.get("records")])

        definitions = workbook.create_sheet("Definitions")
        self._append_rows(
            definitions,
            ["term", "definition"],
            [
                {
                    "term": "Declared ref",
                    "definition": (
                        "A non-empty relationship reference selected in an endpoint payload."
                    ),
                },
                {
                    "term": "Seen ref",
                    "definition": "A declared reference found in the expected endpoint snapshot.",
                },
                {
                    "term": "Missing ref",
                    "definition": (
                        "A valid-looking Centric ref that is not present in the expected "
                        "endpoint snapshot."
                    ),
                },
                {
                    "term": "Invalid ref",
                    "definition": "A non-empty value that does not look like a Centric object ref.",
                },
                {
                    "term": "Endpoint coverage",
                    "definition": (
                        "Current record counts and relationship coverage across the DuckDB store."
                    ),
                },
            ],
        )
        self._format_workbook(workbook)
        workbook.save(path)

    def _normalize_run(self, run: ValidationRunResult | dict[str, Any]) -> dict[str, Any]:
        if isinstance(run, dict):
            return run

        issue_counts = self._legacy_issue_rows(run.results)
        return {
            "rule_set_version": run.rule_set_version,
            "total_products": run.total_products,
            "ready_products": run.ready_products,
            "readiness_percent": run.readiness_percent,
            "summary": {
                "styles": run.total_products,
                "endpoints": 0,
                "relationships_checked": 0,
                "declared_refs": 0,
                "seen_refs": 0,
                "missing_refs": 0,
                "invalid_refs": 0,
                "coverage_percent": run.readiness_percent,
            },
            "relationship_coverage": [],
            "endpoint_coverage": [],
            "unresolved_refs": self._legacy_unresolved_ref_rows(run.results),
            "issue_counts": issue_counts,
        }

    def _summary_rows(self, report: dict[str, Any]) -> list[dict[str, object]]:
        summary = report.get("summary", {})
        return [
            {"metric": "rule_set", "value": report.get("rule_set_version", "")},
            {"metric": "styles_in_store", "value": summary.get("styles", 0)},
            {"metric": "endpoints_in_store", "value": summary.get("endpoints", 0)},
            {
                "metric": "relationships_checked",
                "value": summary.get("relationships_checked", 0),
            },
            {"metric": "declared_refs", "value": summary.get("declared_refs", 0)},
            {"metric": "seen_refs", "value": summary.get("seen_refs", 0)},
            {"metric": "missing_refs", "value": summary.get("missing_refs", 0)},
            {"metric": "invalid_refs", "value": summary.get("invalid_refs", 0)},
            {
                "metric": "reference_coverage_percent",
                "value": summary.get("coverage_percent", 0.0),
            },
        ]

    def _legacy_issue_rows(
        self,
        results: list[ProductValidationResult],
    ) -> list[dict[str, object]]:
        counts: Counter[tuple[str, str, str]] = Counter()
        for result in results:
            for issue in result.issues:
                counts[(issue.code, issue.severity, issue.source_field)] += 1
        return [
            {
                "code": code,
                "severity": severity,
                "bucket": bucket,
                "count": count,
            }
            for (code, severity, bucket), count in sorted(counts.items())
        ]

    def _legacy_unresolved_ref_rows(
        self,
        results: list[ProductValidationResult],
    ) -> list[dict[str, object]]:
        counts: Counter[tuple[str, str, str]] = Counter()
        for result in results:
            for issue in result.issues:
                if issue.source_field == "unresolved_refs" or issue.code in {
                    "INVALID_REFERENCE_VALUE",
                    "REFERENCED_RECORD_NOT_SEEN",
                    "UNRESOLVED_RECONSTRUCTION_REF",
                }:
                    counts[(issue.source_field, issue.source_field, issue.code)] += 1
        return [
            {
                "relationship": relationship,
                "target_endpoint": target_endpoint,
                "status": status,
                "count": count,
            }
            for (relationship, target_endpoint, status), count in sorted(counts.items())
        ]

    def _append_rows(self, sheet, headers, rows) -> None:
        sheet.append(list(headers))
        for row in rows:
            sheet.append([row.get(header) for header in headers])

    def _format_workbook(self, workbook: Workbook) -> None:
        for sheet in workbook.worksheets:
            self._format_sheet(sheet)

    def _format_sheet(self, sheet) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(bottom=Side(style="hair", color="D9E2F3"))
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=cell.column >= 4)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_index in range(1, sheet.max_column + 1):
            letter = get_column_letter(column_index)
            width = self._column_width(sheet, column_index)
            sheet.column_dimensions[letter].width = width

    def _column_width(self, sheet, column_index: int) -> int:
        values = [
            str(sheet.cell(row=row_index, column=column_index).value or "")
            for row_index in range(1, min(sheet.max_row, 200) + 1)
        ]
        max_length = max((len(value) for value in values), default=10)
        return min(max(max_length + 2, 12), 70)
