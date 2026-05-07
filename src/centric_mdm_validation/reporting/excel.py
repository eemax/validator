from __future__ import annotations

from collections.abc import Iterable, Mapping
from dataclasses import dataclass, field
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
FAIL_FILL = "FCE4D6"
HARD_WARNING_FILL = "FFF2CC"
SOFT_WARNING_FILL = "FFF8E1"
PASS_FILL = "E2F0D9"
MID_FILL = "FFF2CC"
LOW_FILL = "FCE4D6"
GRID_COLOR = "D9E2F3"
DEFAULT_HEADER_LABELS = {
    "active_styles": "Active Styles",
    "affected_count": "Affected Count",
    "affected_styles": "Affected Styles",
    "brand": "Brand",
    "bucket": "Bucket",
    "category": "Category",
    "code": "Code",
    "completion": "Completion %",
    "count": "Count",
    "coverage_percent": "Coverage %",
    "declared_refs": "Declared Refs",
    "definition": "Definition",
    "details": "Details",
    "endpoint": "Endpoint",
    "example": "Example",
    "example_message": "Example Message",
    "failed": "Failed",
    "hard_warnings": "Hard Warnings",
    "hydrated_refs": "Seen Refs",
    "invalid_refs": "Invalid Refs",
    "issue": "Issue",
    "issue_count": "Issue Count",
    "issue_type": "Issue Type",
    "metric": "Metric",
    "missing_integration_phase": "Missing Integration Phase",
    "missing_percent": "Missing %",
    "missing_refs": "Missing Refs",
    "passed": "Passed",
    "priority": "Priority",
    "records": "Records",
    "reference_type": "Reference Type",
    "relationship": "Relationship",
    "report_code": "Technical Code",
    "season": "Season",
    "seen_refs": "Seen Refs",
    "severity": "Severity",
    "soft_warnings": "Soft Warnings",
    "source_endpoint": "Source Endpoint",
    "source_records": "Source Records",
    "source_records_with_missing_refs": "Source Records With Missing Refs",
    "source_records_with_refs": "Source Records With Refs",
    "status": "Status",
    "style_id": "Style ID",
    "style_name": "Style Name",
    "style_refs": "Total Refs",
    "styles_with_integration_phase": "Styles With Integration Phase",
    "styles_with_refs": "Styles With Refs",
    "target_endpoint": "Target Endpoint",
    "technical_code": "Technical Code",
    "technical_message": "Technical Message",
    "term": "Term",
    "threshold_percent": "Threshold %",
    "value": "Value",
    "warning_code": "Technical Code",
    "warning_level": "Warning Level",
    "what_needs_fixing": "Action Needed",
}


@dataclass(frozen=True)
class ExcelFormatConfig:
    header_labels: Mapping[str, str] = field(default_factory=lambda: DEFAULT_HEADER_LABELS)
    hidden_headers: frozenset[str] = frozenset(
        {
            "bucket",
            "code",
            "report_code",
            "severity",
            "technical_code",
            "technical_message",
            "warning_code",
            "warning_level",
        }
    )
    wrap_headers: frozenset[str] = frozenset(
        {
            "details",
            "example",
            "example_message",
            "issue_codes",
            "report_code",
            "technical_message",
            "warning_code",
            "warning_codes",
            "what_needs_fixing",
        }
    )
    percent_headers: frozenset[str] = frozenset(
        {
            "completion",
            "completion_percent",
            "coverage_percent",
            "missing_percent",
            "readiness_percent",
            "threshold_percent",
        }
    )
    integer_hints: tuple[str, ...] = (
        "styles",
        "passed",
        "failed",
        "warnings",
        "count",
        "refs",
        "records",
        "hydrated",
        "missing",
        "rows",
        "boms",
        "materials",
        "suppliers",
        "factories",
        "value",
    )


DEFAULT_FORMAT_CONFIG = ExcelFormatConfig()


def append_rows(sheet, headers: Iterable[str], rows: Iterable[Any], *, config=None) -> None:
    config = config or DEFAULT_FORMAT_CONFIG
    header_list = list(headers)
    sheet.append([_display_header(header, config) for header in header_list])
    for row in rows:
        if isinstance(row, dict):
            sheet.append([row.get(header) for header in header_list])
        else:
            sheet.append(list(row))


def format_workbook(workbook, *, config=None) -> None:
    config = config or DEFAULT_FORMAT_CONFIG
    for sheet in workbook.worksheets:
        format_sheet(sheet, config=config)


def format_sheet(sheet, *, config=None) -> None:
    config = config or DEFAULT_FORMAT_CONFIG
    if sheet.max_row < 1:
        return
    _style_header(sheet)
    _style_body(sheet, config)
    _set_dimensions(sheet, config)
    _set_filters_and_freeze(sheet)


def _style_header(sheet) -> None:
    header_fill = PatternFill("solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_FONT)
    border = Border(bottom=Side(style="thin", color=GRID_COLOR))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_body(sheet, config: ExcelFormatConfig) -> None:
    thin_border = Border(bottom=Side(style="hair", color=GRID_COLOR))
    headers = _headers(sheet, config)
    for row in sheet.iter_rows(min_row=2):
        fill = _row_fill(headers, row)
        for cell in row:
            header = headers.get(cell.column)
            cell.border = thin_border
            cell.alignment = _alignment_for_header(header, config)
            if fill is not None:
                cell.fill = fill
            _apply_number_format(cell, header, config)


def _set_dimensions(sheet, config: ExcelFormatConfig) -> None:
    sheet.row_dimensions[1].height = 28
    for column_index in range(1, sheet.max_column + 1):
        letter = get_column_letter(column_index)
        header = str(sheet.cell(row=1, column=column_index).value or "")
        key = _header_key(header, config)
        width = _column_width(sheet, column_index, key, config)
        sheet.column_dimensions[letter].width = width
        if key in config.hidden_headers:
            sheet.column_dimensions[letter].hidden = True
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 30


def _set_filters_and_freeze(sheet) -> None:
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions


def _column_width(sheet, column_index, header, config: ExcelFormatConfig) -> float:
    display_header = str(sheet.cell(row=1, column=column_index).value or header)
    max_length = len(display_header)
    sample_limit = min(sheet.max_row, 250)
    for row_index in range(2, sample_limit + 1):
        value = sheet.cell(row=row_index, column=column_index).value
        if value is None:
            continue
        max_length = max(max_length, len(str(value)))
    if header in config.wrap_headers:
        return min(max(max_length * 0.75, 24), 70)
    return min(max(max_length + 2, 10), 38)


def _headers(sheet, config: ExcelFormatConfig) -> dict[int, str]:
    return {cell.column: _header_key(str(cell.value or ""), config) for cell in sheet[1]}


def _alignment_for_header(header, config: ExcelFormatConfig) -> Alignment:
    if header in config.wrap_headers:
        return Alignment(wrap_text=True, vertical="top")
    if _is_numeric_header(header, config):
        return Alignment(horizontal="right", vertical="top")
    return Alignment(vertical="top")


def _apply_number_format(cell, header, config: ExcelFormatConfig) -> None:
    if cell.value is None:
        return
    if header in config.percent_headers:
        cell.number_format = '0.00"%"'
        return
    if isinstance(cell.value, int) or (
        isinstance(cell.value, float) and _is_numeric_header(header, config)
    ):
        cell.number_format = "#,##0"


def _row_fill(headers, row):
    row_values = {headers.get(cell.column): cell.value for cell in row}
    status = str(row_values.get("status") or "").lower()
    if status == "failed":
        return PatternFill("solid", fgColor=FAIL_FILL)
    warning_level = str(row_values.get("warning_level") or "").lower()
    if warning_level == "hard":
        return PatternFill("solid", fgColor=HARD_WARNING_FILL)
    if warning_level == "soft":
        return PatternFill("solid", fgColor=SOFT_WARNING_FILL)
    priority = str(row_values.get("priority") or "").lower()
    if priority == "hard warning":
        return PatternFill("solid", fgColor=HARD_WARNING_FILL)
    if priority == "soft warning":
        return PatternFill("solid", fgColor=SOFT_WARNING_FILL)
    if priority == "blocking":
        return PatternFill("solid", fgColor=FAIL_FILL)
    if (row_values.get("failed") or 0) > 0:
        return PatternFill("solid", fgColor=FAIL_FILL)
    if (row_values.get("hard_warnings") or 0) > 0:
        return PatternFill("solid", fgColor=HARD_WARNING_FILL)
    completion = row_values.get("completion")
    if isinstance(completion, (int, float)):
        if completion >= 90:
            return PatternFill("solid", fgColor=PASS_FILL)
        if completion >= 50:
            return PatternFill("solid", fgColor=MID_FILL)
        return PatternFill("solid", fgColor=LOW_FILL)
    return None


def _is_numeric_header(header, config: ExcelFormatConfig) -> bool:
    return header in config.percent_headers or any(hint in header for hint in config.integer_hints)


def _display_header(header, config: ExcelFormatConfig) -> str:
    return config.header_labels.get(header, str(header).replace("_", " ").title())


def _header_key(header, config: ExcelFormatConfig) -> str:
    display_header_keys = {
        label: key for key, label in config.header_labels.items() if label not in {"Technical Code"}
    }
    display_header_keys.update(
        {
            "Technical Code": "technical_code",
            "Completion %": "completion",
            "Coverage %": "coverage_percent",
            "Missing %": "missing_percent",
            "Threshold %": "threshold_percent",
        }
    )
    if header in display_header_keys:
        return display_header_keys[header]
    return str(header).strip().lower().replace("%", "percent").replace(" ", "_")

