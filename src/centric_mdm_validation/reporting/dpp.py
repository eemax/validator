import csv
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook

from centric_mdm_validation.models import ProductValidationResult, ValidationRunResult


class DppReadinessReporter:
    def write_all(self, run: ValidationRunResult, output_dir: Path) -> None:
        output_dir.mkdir(parents=True, exist_ok=True)
        self.write_summary_markdown(run, output_dir / "dpp-readiness-summary.md")
        self.write_product_csv(run, output_dir / "dpp-readiness-products.csv")
        self.write_issue_csv(run, output_dir / "dpp-readiness-issues.csv")
        self.write_workbook(run, output_dir / "dpp-readiness.xlsx")

    def write_summary_markdown(self, run: ValidationRunResult, path: Path) -> None:
        brand_rows = self._brand_rows(run.results)
        missing_rows = self._missing_attribute_rows(run.results)

        lines = [
            "# DPP Readiness Summary",
            "",
            f"- Rule set: `{run.rule_set_version}`",
            f"- Products checked: `{run.total_products}`",
            f"- Ready products: `{run.ready_products}`",
            f"- Readiness: `{run.readiness_percent}%`",
            "",
            "## Readiness By Brand",
            "",
            "| Brand | Products | Ready | Readiness | Avg Score |",
            "| --- | ---: | ---: | ---: | ---: |",
        ]
        lines.extend(
            f"| {row['brand']} | {row['products']} | {row['ready']} | "
            f"{row['readiness_percent']}% | {row['average_score']} |"
            for row in brand_rows
        )
        lines.extend(
            [
                "",
                "## Most Common Missing Attributes",
                "",
                "| Attribute | Count |",
                "| --- | ---: |",
            ]
        )
        lines.extend(f"| {row['attribute']} | {row['count']} |" for row in missing_rows)
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    def write_product_csv(self, run: ValidationRunResult, path: Path) -> None:
        with path.open("w", encoding="utf-8", newline="") as file:
            writer = csv.DictWriter(
                file,
                fieldnames=[
                    "centric_style_id",
                    "brand_code",
                    "brand_name",
                    "product_type_code",
                    "ready",
                    "score",
                    "issue_count",
                    "blocking_issue_count",
                    "warning_count",
                ],
            )
            writer.writeheader()
            for result in run.results:
                writer.writerow(
                    {
                        "centric_style_id": result.centric_style_id,
                        "brand_code": result.brand_code,
                        "brand_name": result.brand_name,
                        "product_type_code": result.product_type_code,
                        "ready": result.ready,
                        "score": result.score,
                        "issue_count": result.issue_count,
                        "blocking_issue_count": result.blocking_issue_count,
                        "warning_count": result.warning_count,
                    }
                )

    def write_issue_csv(self, run: ValidationRunResult, path: Path) -> None:
        with path.open("w", encoding="utf-8", newline="") as file:
            writer = csv.DictWriter(
                file,
                fieldnames=[
                    "centric_style_id",
                    "brand_code",
                    "product_type_code",
                    "code",
                    "severity",
                    "blocking",
                    "fix_location",
                    "source_field",
                    "rule_id",
                    "message",
                ],
            )
            writer.writeheader()
            for result in run.results:
                for issue in result.issues:
                    writer.writerow(
                        {
                            "centric_style_id": result.centric_style_id,
                            "brand_code": result.brand_code,
                            "product_type_code": result.product_type_code,
                            "code": issue.code,
                            "severity": issue.severity,
                            "blocking": issue.blocking,
                            "fix_location": issue.fix_location,
                            "source_field": issue.source_field,
                            "rule_id": issue.rule_id,
                            "message": issue.message,
                        }
                    )

    def write_workbook(self, run: ValidationRunResult, path: Path) -> None:
        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append(["Metric", "Value"])
        summary.append(["Rule set", run.rule_set_version])
        summary.append(["Products checked", run.total_products])
        summary.append(["Ready products", run.ready_products])
        summary.append(["Readiness percent", run.readiness_percent])

        by_brand = workbook.create_sheet("By Brand")
        by_brand.append(["Brand", "Products", "Ready", "Readiness %", "Average Score"])
        for row in self._brand_rows(run.results):
            by_brand.append(
                [
                    row["brand"],
                    row["products"],
                    row["ready"],
                    row["readiness_percent"],
                    row["average_score"],
                ]
            )

        products = workbook.create_sheet("Products")
        products.append(
            [
                "Centric Style ID",
                "Brand Code",
                "Brand Name",
                "Product Type",
                "Ready",
                "Score",
                "Issues",
                "Blocking Issues",
                "Warnings",
            ]
        )
        for result in run.results:
            products.append(
                [
                    result.centric_style_id,
                    result.brand_code,
                    result.brand_name,
                    result.product_type_code,
                    result.ready,
                    result.score,
                    result.issue_count,
                    result.blocking_issue_count,
                    result.warning_count,
                ]
            )

        issues = workbook.create_sheet("Issues")
        issues.append(
            [
                "Centric Style ID",
                "Brand",
                "Product Type",
                "Code",
                "Severity",
                "Blocking",
                "Fix Location",
                "Source Field",
                "Rule ID",
                "Message",
            ]
        )
        for result in run.results:
            for issue in result.issues:
                issues.append(
                    [
                        result.centric_style_id,
                        result.brand_code,
                        result.product_type_code,
                        issue.code,
                        issue.severity,
                        issue.blocking,
                        issue.fix_location,
                        issue.source_field,
                        issue.rule_id,
                        issue.message,
                    ]
                )

        workbook.save(path)

    def _brand_rows(self, results: list[ProductValidationResult]) -> list[dict[str, object]]:
        grouped: dict[str, list[ProductValidationResult]] = defaultdict(list)
        for result in results:
            grouped[result.brand_code or "UNKNOWN"].append(result)

        rows = []
        for brand, brand_results in sorted(grouped.items()):
            products = len(brand_results)
            ready = sum(1 for result in brand_results if result.ready)
            average_score = round(sum(result.score for result in brand_results) / products, 2)
            rows.append(
                {
                    "brand": brand,
                    "products": products,
                    "ready": ready,
                    "readiness_percent": round((ready / products) * 100, 2),
                    "average_score": average_score,
                }
            )
        return rows

    def _missing_attribute_rows(
        self,
        results: list[ProductValidationResult],
    ) -> list[dict[str, object]]:
        counts: Counter[str] = Counter()
        for result in results:
            for issue in result.issues:
                if issue.code in {
                    "DPP_REQUIRED_ATTRIBUTE_MISSING",
                    "DPP_RECOMMENDED_ATTRIBUTE_MISSING",
                }:
                    counts[issue.source_field.removeprefix("attributes.")] += 1
        return [
            {"attribute": attribute, "count": count}
            for attribute, count in counts.most_common()
        ]
